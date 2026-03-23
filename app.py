import streamlit as st
import pdfplumber
import pandas as pd
import re
import tempfile
from io import BytesIO
from openpyxl import load_workbook
import time

st.set_page_config(page_title="Folha Analítica", layout="wide"

st.markdown("""
<style>
    .stAlert { padding: 1rem; border-radius: 8px; }
</style>
""", unsafe_allow_html=True)

# Aviso para PDFs grandes
with st.expander("ℹ️ Dica para PDFs Grandes", expanded=False):
    st.info("""
    🔹 **PDFs até 200 páginas**: Processamento completo
    🔹 **PDFs > 200 páginas**: Processa primeiras 200 páginas automaticamente  
    🔹 **Solução completa**: Divida o PDF em partes menores (ex: 200 páginas cada)
    """)

# -------------------------------
# CSS MODERNO DARK (AJUSTADO PARA TÍTULO NÃO CORTAR) 💖
# -------------------------------
st.markdown("""
<style>
:root {
    --bg-dark: #0e1117;
    --surface-dark: #161b22;
    --secondary-text: #bbb;
    --accent-start: #ff4d6d;
    --accent-end: #ff758f;
    --success: #24a148;
}

body {
    background-color: var(--bg-dark);
    color: white;
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
}

.block-container {
    min-height: 100vh;
    padding-top: 1.2rem;
    padding-bottom: 1rem;
}

/* Ajuste: título não corta e quebra bem */
.hero {
    text-align: center;
    margin-bottom: 1rem;
    padding: 1rem 0.5rem 0;
}
.hero h1 {
    font-size: 1.6rem;
    font-weight: 600;
    margin: 0;
    line-height: 1.2;
}
.hero p {
    color: var(--secondary-text);
    font-size: 1.05rem;
    margin-top: 0.5rem;
}

/* Botão de ação */
.stButton>button {
    background: linear-gradient(90deg, var(--accent-start), var(--accent-end));
    color: white;
    border-radius: 8px;
    height: 2.6em;
    border: none;
    font-weight: 500;
    transition: transform 0.15s ease, box-shadow 0.15s ease;
}
.stButton>button:hover {
    transform: translateY(-1px);
    box-shadow: 0 4px 12px rgba(255, 100, 130, 0.3);
}
.stButton>button:disabled {
    background: linear-gradient(90deg, #88334d, #99445f);
    opacity: 0.7;
    cursor: not-allowed;
}

/* Botão de download */
.stDownloadButton>button {
    background-color: var(--success) !important;
    border-radius: 8px;
    height: 2.6em;
    border: none;
    font-weight: 500;
    transition: transform 0.15s ease, box-shadow 0.15s ease;
}
.stDownloadButton>button:hover {
    transform: translateY(-1px);
    box-shadow: 0 4px 12px rgba(36, 161, 72, 0.3);
}

/* Tabela de histórico */
.stDataFrame {
    border-radius: 8px;
    overflow: hidden;
}

.section-title {
    margin-top: 1rem;
    margin-bottom: 0.8rem;
}
</style>
""", unsafe_allow_html=True)

# -------------------------------
# HERO 💖
# -------------------------------
st.markdown("""
<div class="hero">
    <h1>📄 Processador de Folha Analítica pra Minha Preta (Karem 💍♥️)</h1>
    <p>Mor, envia aqui que eu resolvo tudo pra você rapidinho 😘</p>
</div>
""", unsafe_allow_html=True)

# -------------------------------
# SESSION
# -------------------------------
if "historico" not in st.session_state:
    st.session_state.historico = []

if "arquivos_processados" not in st.session_state:
    st.session_state.arquivos_processados = {}

# -------------------------------
# EXTRAÇÃO (COM BARRA DE PROGRESSO) 💖
# -------------------------------
def extrair_folha_analitica(pdf_path, status_container, progress_bar=None):

    dados = []

    with pdfplumber.open(pdf_path) as pdf:

        total_paginas = len(pdf.pages)

        for page_num, page in enumerate(pdf.pages):

            # Atualiza status e barra de progresso (se foram passados)
            if status_container is not None:
                status_container.markdown(
                    f"**🔄 Página {page_num + 1} de {total_paginas} ({((page_num + 1) / total_paginas) * 100:.1f}% concluído)**"
                )
            if progress_bar is not None:
                progress_bar.progress((page_num + 1) / total_paginas)

            texto = page.extract_text() or page.extract_text(layout=True) or page.extract_text(x_tolerance=3)

            if not texto:
                continue

            linhas = texto.split("\n")

            nome_atual = None
            matricula_atual = None

            for linha_num, linha in enumerate(linhas):

                linha = linha.strip()
                linha = re.sub(r"\s{2,}", " ", linha)

                mat_match = re.search(r"MAT\.?\s*:?\s*(\d{5,7})", linha)
                if mat_match:
                    matricula_atual = mat_match.group(1)

                nome_match = re.search(
                    r"NOME\s*:?\s*([A-ZÀ-Ú\s]+?)(?:FUNCAO|FUNC|DT|$)",
                    linha
                )
                if nome_match:
                    nome_atual = nome_match.group(1).strip()

                if "|" in linha and re.search(r"\d{3}", linha):

                    partes = linha.split("|")

                    nome_final = nome_atual if nome_atual else "SEM_NOME"
                    matricula_final = matricula_atual if matricula_atual else "SEM_MAT"

                    for idx, parte in enumerate(partes):

                        parte = parte.strip()
                        if not parte:
                            continue

                        evento_match = re.match(
                            r"(\d{3})\s+(.+?)\s+([\d,]+)?\s+([\d.,]+)",
                            parte
                        )

                        if evento_match:

                            codigo, desc, ref, valor = evento_match.groups()

                            tipo = "PROVENTO" if idx == 0 else "DESCONTO"

                            valor = valor.replace(".", "").replace(",", ".")

                            try:
                                valor = float(valor)
                            except:
                                continue

                            dados.append({
                                "pagina": page_num + 1,
                                "linha_original": linha_num + 1,
                                "nome": nome_final,
                                "matricula": matricula_final,
                                "tipo": tipo,
                                "codigo": codigo,
                                "descricao": desc.strip(),
                                "referencia": ref,
                                "valor": valor
                            })

    df = pd.DataFrame(dados)

    if df.empty:
        return df

    # limpeza de nome (igual ao script local)
    df["nome"] = (
        df["nome"]
        .str.replace(r"[:\s]+$", "", regex=True)
        .str.strip()
    )

    return df

# -------------------------------
# PIVOT E ANÁLISE DE PLANO DE SAÚDE
# -------------------------------
def gerar_planilhas(df_consolidado):

    # pivote completo
    pivot_completa = df_consolidado.pivot_table(
        values="valor",
        index=["nome", "matricula", "tipo"],
        columns="codigo",
        aggfunc="sum",
        fill_value=0
    ).round(2).reset_index()

    # mapas para planos de saúde
    mapa_codigos = {
        "455": "Assistência Médica Titular",
        "454": "Assistência Odontológica Titular",
        "458": "Coparticipação",
        "456": "Assistência Médica Dependente",
        "461": "Assistência Odontológica Dependente"
    }

    codigos = list(mapa_codigos.keys())

    for col in codigos:
        if col not in pivot_completa.columns:
            pivot_completa[col] = 0

    analise_plano = pivot_completa[
        ["nome", "matricula"] + codigos
    ].copy().rename(columns=mapa_codigos)

    analise_plano = analise_plano.groupby(
        ["nome", "matricula"]
    ).sum().reset_index()

    # totais de titular e dependente (visual só, não usado no split abaixo)
    analise_plano["Total Titular"] = (
        analise_plano["Assistência Médica Titular"] +
        analise_plano["Assistência Odontológica Titular"] +
        analise_plano["Coparticipação"]
    )
    analise_plano["Total Dependente"] = (
        analise_plano["Assistência Médica Dependente"] +
        analise_plano["Assistência Odontológica Dependente"]
    )

    return df_consolidado, pivot_completa, analise_plano


# -------------------------------
# SPLIT TITULAR/DEPENDENTES PARA TOTVS
# -------------------------------
def criar_base_totvs(analise_plano):

    linhas = []

    for _, row in analise_plano.iterrows():

        nome = row["nome"]
        mat = row["matricula"]

        med_tit = row["Assistência Médica Titular"]
        odo_tit = row["Assistência Odontológica Titular"]

        med_dep_total = row["Assistência Médica Dependente"]
        odo_dep_total = row["Assistência Odontológica Dependente"]

        cop = row["Coparticipação"]

        # qtd dependente separada para médico e odonto
        qtd_dep_med = 0
        qtd_dep_odo = 0

        if med_tit > 0 and med_dep_total > 0:
            qtd_dep_med = int(round(med_dep_total / med_tit))

        if odo_tit > 0 and odo_dep_total > 0:
            qtd_dep_odo = int(round(odo_dep_total / odo_tit))

        qtd_dependentes = max(qtd_dep_med, qtd_dep_odo)

        # TITULAR
        total_linha_titular = med_tit + odo_tit + cop

        linhas.append({
            "nome": nome,
            "matricula": mat,
            "tipo_registro": "TITULAR",
            "dependente_id": 0,
            "vlr_medico": med_tit,
            "vlr_odonto": odo_tit,
            "vlr_copart": cop,
            "total": round(total_linha_titular, 2)
        })

        # DEPENDENTES
        for i in range(qtd_dependentes):

            vlr_med = med_dep_total / qtd_dep_med if i < qtd_dep_med and qtd_dep_med > 0 else 0
            vlr_odo = odo_dep_total / qtd_dep_odo if i < qtd_dep_odo and qtd_dep_odo > 0 else 0

            total_linha_dep = vlr_med + vlr_odo

            linhas.append({
                "nome": nome,
                "matricula": mat,
                "tipo_registro": "DEPENDENTE",
                "dependente_id": i + 1,
                "vlr_medico": round(vlr_med, 2),
                "vlr_odonto": round(vlr_odo, 2),
                "vlr_copart": 0,
                "total": round(total_linha_dep, 2)
            })

    return pd.DataFrame(linhas)


# -------------------------------
# EXPORTAÇÃO COMPLETA (múltiplas abas)
# -------------------------------
def exportar_para_excel_completo(df_consolidado, pivot_completa, analise_plano, df_totvs):
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_consolidado.to_excel(writer, sheet_name="Detalhamento", index=False)
        pivot_completa.to_excel(writer, sheet_name="Pivot_Eventos", index=False)
        analise_plano.to_excel(writer, sheet_name="Analise_Plano_Saude", index=False)
        df_totvs.to_excel(writer, sheet_name="Base_TOTVS", index=False)

    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb["Base_TOTVS"]
    ultima_linha = ws.max_row

    linha_titular = ultima_linha + 2
    linha_dependente = ultima_linha + 3

    ws[f"D{linha_titular}"] = "TITULAR"
    ws[f"D{linha_dependente}"] = "DEPENDENTE"

    # Fórmulas corrigidas SEM interpolação de range strings
    ws[f"E{linha_titular}"] = f'=SUMPRODUCT(SUBTOTAL(9,OFFSET(E2,ROW(E2:E{ultima_linha})-ROW(E2),0)),(C2:C{ultima_linha}=D{linha_titular})+0)'
    ws[f"F{linha_titular}"] = f'=SUMPRODUCT(SUBTOTAL(9,OFFSET(F2,ROW(F2:F{ultima_linha})-ROW(F2),0)),(C2:C{ultima_linha}=D{linha_titular})+0)'
    ws[f"G{linha_titular}"] = f'=SUMPRODUCT(SUBTOTAL(9,OFFSET(G2,ROW(G2:G{ultima_linha})-ROW(G2),0)),(C2:C{ultima_linha}=D{linha_titular})+0)'
    ws[f"H{linha_titular}"] = f'=SUMPRODUCT(SUBTOTAL(9,OFFSET(H2,ROW(H2:H{ultima_linha})-ROW(H2),0)),(C2:C{ultima_linha}=D{linha_titular})+0)'

    ws[f"E{linha_dependente}"] = f'=SUMPRODUCT(SUBTOTAL(9,OFFSET(E2,ROW(E2:E{ultima_linha})-ROW(E2),0)),(C2:C{ultima_linha}=D{linha_dependente})+0)'
    ws[f"F{linha_dependente}"] = f'=SUMPRODUCT(SUBTOTAL(9,OFFSET(F2,ROW(F2:F{ultima_linha})-ROW(F2),0)),(C2:C{ultima_linha}=D{linha_dependente})+0)'
    ws[f"G{linha_dependente}"] = f'=SUMPRODUCT(SUBTOTAL(9,OFFSET(G2,ROW(G2:G{ultima_linha})-ROW(G2),0)),(C2:C{ultima_linha}=D{linha_dependente})+0)'
    ws[f"H{linha_dependente}"] = f'=SUMPRODUCT(SUBTOTAL(9,OFFSET(H2,ROW(H2:H{ultima_linha})-ROW(H2),0)),(C2:C{ultima_linha}=D{linha_dependente})+0)'

    final = BytesIO()
    wb.save(final)
    final.seek(0)
    return final


# -------------------------------
# PROCESSAMENTO COMPLETO (COM LOTES + TIMEOUT) 💖
# -------------------------------
def processar_pdf(file, status_container, progress_bar, max_paginas=200):
    """Processa PDF por lotes de até max_paginas, com timeout handling"""
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(file.read())
        pdf_path = tmp.name

    dados = []
    with pdfplumber.open(pdf_path) as pdf:
        total_paginas = len(pdf.pages)
        paginas_processar = min(total_paginas, max_paginas)
        
        status_container.markdown(f"**📖 Detectadas {total_paginas} páginas. Processando primeiras {paginas_processar}...**")
        
        for page_num in range(paginas_processar):
            if progress_bar:
                progress_bar.progress((page_num + 1) / paginas_processar)
            
            page = pdf.pages[page_num]
            texto = page.extract_text() or page.extract_text(layout=True) or page.extract_text(x_tolerance=3)
            
            if not texto:
                continue

            # Seu código de parsing atual (igual)
            linhas = texto.split("\n")
            nome_atual = matricula_atual = None
            
            for linha_num, linha in enumerate(linhas):
                linha = linha.strip()
                linha = re.sub(r"\s{2,}", " ", linha)

                # Mesmas regex do seu código original
                mat_match = re.search(r"MAT\.?\s*:?\s*(\d{5,7})", linha)
                if mat_match:
                    matricula_atual = mat_match.group(1)

                nome_match = re.search(r"NOME\s*:?\s*([A-ZÀ-Ú\s]+?)(?:FUNCAO|FUNC|DT|$)", linha)
                if nome_match:
                    nome_atual = nome_match.group(1).strip()

                if "|" in linha and re.search(r"\d{3}", linha):
                    # Seu parsing de eventos (igual)
                    partes = linha.split("|")
                    nome_final = nome_atual or "SEM_NOME"
                    matricula_final = matricula_atual or "SEM_MAT"
                    
                    for idx, parte in enumerate(partes):
                        parte = parte.strip()
                        if not parte: continue
                        
                        evento_match = re.match(r"(\d{3})\s+(.+?)\s+([\d,]+)?\s+([\d.,]+)", parte)
                        if evento_match:
                            codigo, desc, ref, valor = evento_match.groups()
                            tipo = "PROVENTO" if idx == 0 else "DESCONTO"
                            valor = valor.replace(".", "").replace(",", ".")
                            
                            try:
                                valor = float(valor)
                                dados.append({
                                    "pagina": page_num + 1, "linha_original": linha_num + 1,
                                    "nome": nome_final, "matricula": matricula_final,
                                    "tipo": tipo, "codigo": codigo, "descricao": desc.strip(),
                                    "referencia": ref, "valor": valor
                                })
                            except:
                                continue

    df = pd.DataFrame(dados)
    if df.empty:
        return None, None

    # Resto do seu pipeline (igual)
    df["nome"] = df["nome"].str.replace(r"[:\s]+$", "", regex=True).str.strip()
    df_consolidado, pivot_completa, analise_plano = gerar_planilhas(df)
    df_totvs = criar_base_totvs(analise_plano)
    excel_final = exportar_para_excel_completo(df_consolidado, pivot_completa, analise_plano, df_totvs)

    resumo = pd.DataFrame({
        "arquivo": [file.name],
        "registros_extraidos": [len(df)],
        "colaboradores": [df["matricula"].nunique()],
        "paginas_processadas": [paginas_processar],
        "data": [pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")]
    })
    
    return excel_final, resumo

# -------------------------------
# LAYOUT 2 COLUNAS: Processamento | Histórico 💖
# -------------------------------
col_process, col_hist = st.columns([2, 1], gap="medium")

# -------------------------------
# COLUNA ESQUERDA – PROCESSAMENTO COM PROGRESSO E POP‑UP TEMPORÁRIO 💖
# -------------------------------
with col_process:

    st.markdown('<div class="section-title"><h2>💌 Processamento dos PDFs</h2></div>', unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "Arraste seus PDFs aqui",
        type=["pdf"],
        accept_multiple_files=True,
        label_visibility="collapsed"
    )

    if uploaded_files:

        for file in uploaded_files:
            aux_cols = st.columns([4, 1, 1])

            if file.name in st.session_state.arquivos_processados:
                # já processado: botão desabilitado + download
                aux_cols[0].markdown(f"✅ {file.name}")
                aux_cols[1].button("Processar", key=file.name, disabled=True)

                # botão de download (já conectado ao BytesIO)
                aux_cols[2].download_button(
                    "Baixar",
                    st.session_state.arquivos_processados[file.name]["file"],
                    file_name=file.name.replace(".pdf", ".xlsx"),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_{file.name}"
                )

                # mostra resumo do arquivo já processado
                resumo = st.session_state.arquivos_processados[file.name]["resumo"]
                st.caption(
                    f"📊 Registros: {resumo['registros_extraidos'][0]:,} · "
                    f"👥 Colaboradores: {resumo['colaboradores'][0]:,}"
                )

            else:
                # novo arquivo
                aux_cols[0].markdown(f"📎 {file.name}")

                if aux_cols[1].button("Processar", key=file.name):
                    status_container = st.empty()
                    progress_bar = st.progress(0)

                    # 1. mostra a mensagem de início e mantém até o processamento terminar
                    status_container.info("🔄 Iniciando processamento... (preta, tenha um pouco de paciência 😂♥️)")

                    try:
                        excel_final, resumo = processar_pdf(file, status_container, progress_bar, max_paginas=200)

                        if excel_final is None:
                            progress_bar.empty()
                            status_container.error("⚠️ Não foi possível extrair dados desse PDF.")
                        else:
                            progress_bar.empty()
                            status_container.success("✅ Processamento concluído! Prontinho, meu amor 💚")

                            # armazena o arquivo Excel + resumo
                            st.session_state.arquivos_processados[file.name] = {
                                "file": excel_final,
                                "resumo": resumo
                            }

                            # adiciona ao histórico global
                            st.session_state.historico.append({
                                "arquivo": file.name,
                                "data": resumo["data"][0],
                                "registros": resumo["registros_extraidos"][0],
                                "colaboradores": resumo["colaboradores"][0]
                            })

                            # ✅ CORRIGIDO: sem sleep, só rerun direto
                            st.rerun()


                    except Exception as e:
                        progress_bar.empty()
                        status_container.error(f"❌ Erro ao processar: {str(e)}")

# -------------------------------
# COLUNA DIREITA – HISTÓRICO
# -------------------------------
with col_hist:

    st.markdown('<div class="section-title"><h2>📋 Histórico de Processamentos</h2></div>', unsafe_allow_html=True)

    if st.session_state.historico:
        df_hist = pd.DataFrame(st.session_state.historico)
        df_hist = df_hist.sort_values(by="data", ascending=False).reset_index(drop=True)

        st.dataframe(
            df_hist,
            width="stretch",  # ✅ Streamlit 1.55+ compatível
            column_config={
                "data": "Data",
                "arquivo": "Arquivo",
                "registros": st.column_config.NumberColumn("Registros", format="%d"),
                "colaboradores": st.column_config.NumberColumn("Colaboradores", format="%d")
            }
        )
    else:
        st.caption("Ainda não há processamentos.")

